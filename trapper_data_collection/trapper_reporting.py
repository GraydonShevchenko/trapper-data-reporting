import sys, os
import pandas as pd
import boto3
import openpyxl
import datetime as dt
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from arcgis.gis import GIS
from arcgis.geometry import filters
from argparse import ArgumentParser
from collections import defaultdict
import logging

from util.environment import Environment

import trap_config


def run_app():
    ago_user, ago_pass, obj_store_user, obj_store_secret, obj_store_host, logger = get_input_parameters()
    try:
        report = TrapReport(ago_user=ago_user, ago_pass=ago_pass, obj_store_user=obj_store_user, 
                           obj_store_secret=obj_store_secret, obj_store_host=obj_store_host, logger=logger)

        report.download_attachments()
        report.create_excel()
        report.create_wild_report()

        del report
    except Exception as e:
        logger.exception('There was an exception')


def get_input_parameters():
    """
    Function:
        Sets up parameters and the logger object
    Returns:
        tuple: user entered parameters required for tool execution
    """
    try:
        parser = ArgumentParser(description='This script is used to update the Traps AGOL feature layer based on information entered in the trap check table')
        # parser.add_argument('ago_user', nargs='?', type=str, help='AGOL Username')
        # parser.add_argument('ago_pass', nargs='?', type=str, help='AGOL Password')
        parser.add_argument('--log_level', default='INFO', choices=['DEBUG', 'INFO', 'WARNING', 'ERROR'],
                            help='Log level')
        parser.add_argument('--log_dir', help='Path to log directory')

        args = parser.parse_args()
        try:
            ago_user = trap_config.AGO_USER
            ago_pass = trap_config.AGO_PASS
            obj_store_user = trap_config.OBJ_STORE_USER
            obj_store_secret = trap_config.OBJ_STORE_SECRET
            obj_store_host = trap_config.OBJ_STORE_HOST
        except:
            ago_user = os.environ['AGO_USER']
            ago_pass = os.environ['AGO_PASS']
            obj_store_user = os.environ['OBJ_STORE_USER']
            obj_store_secret = os.environ['OBJ_STORE_SECRET']
            obj_store_host = os.environ['OBJ_STORE_HOST']

        logger = Environment.setup_logger(args)

        return ago_user, ago_pass, obj_store_user, obj_store_secret, obj_store_host, logger

    except Exception as e:
        logging.error('Unexpected exception. Program terminating: {}'.format(e.message))
        raise Exception('Errors exist')


class TrapReport:
    def __init__(self, ago_user, ago_pass, obj_store_user, obj_store_secret, obj_store_host, logger) -> None:
        self.ago_user = ago_user
        self.ago_pass = ago_pass
        self.obj_store_user = obj_store_user
        self.obj_store_secret = obj_store_secret
        self.obj_store_host = obj_store_host
        self.logger = logger

        self.portal_url = trap_config.MAPHUB
        self.ago_traps = trap_config.TRAPS
        self.ago_fisher = trap_config.FISHER
        self.ago_parks = trap_config.PARKS
        self.ago_wmu = trap_config.WMU

        self.trapper_bucket = trap_config.BUCKET
        self.bucket_prefix ='trapper_data_collection'

        self.logger.info('Connecting to map hub')
        self.gis = GIS(url=self.portal_url, username=self.ago_user, password=self.ago_pass, expiration=9999)
        self.logger.info('Connection successful')

        self.logger.info('Connecting to object storage')
        self.boto_resource = boto3.resource(service_name='s3', 
                                            aws_access_key_id=self.obj_store_user,
                                            aws_secret_access_key=self.obj_store_secret, 
                                            endpoint_url=f'https://{self.obj_store_host}')

    def __del__(self) -> None:
        self.logger.info('Disconnecting from maphub')
        del self.gis
        self.logger.info('Closing object storage connection')
        del self.boto_resource

    def list_contents(self) -> list:
        obj_bucket = self.boto_resource.Bucket(self.trapper_bucket)
        lst_objects = []
        for obj in obj_bucket.objects.all():
            lst_objects.append(os.path.basename(obj.key))

        return lst_objects


    def download_attachments(self) -> None:
        """
        Function:
            Master function to download attachments for all required layers in arcgis online
        Returns:
            None
        """
        lst_pictures = self.list_contents()

        self.copy_to_object_storage(ago_layer=self.ago_traps, layer_name='traps', 
                                    fld_picture='PICTURE', lst_os_pictures=lst_pictures, folder='trap_setup')
        
        self.copy_to_object_storage(ago_layer=self.ago_traps, layer_name='trap checks', 
                                    fld_picture='PICTURE', lst_os_pictures=lst_pictures, folder='trap_check')
        
        self.copy_to_object_storage(ago_layer=self.ago_fisher, layer_name='fisher', 
                                    fld_picture='PICTURE', lst_os_pictures=lst_pictures, folder='fisher')
        

    def copy_to_object_storage(self, ago_layer, layer_name, fld_picture, lst_os_pictures, folder) -> None:
        """
        Function:
            Function used to download attachments from arcgis online layers and copy them to object storage.
        Returns:
            None
        """
        self.logger.info(f'Downloading photos on the {layer_name} layer')
        ago_item = self.gis.content.get(ago_layer)
        if layer_name != 'trap checks':
            ago_flayer = ago_item.layers[0]
        else:
            ago_flayer = ago_item.tables[0]

        ago_fset = ago_flayer.query()
        all_features = ago_fset.features
        if len(all_features) == 0:
            return

        lst_oids = ago_fset.sdf['OBJECTID'].tolist()

        for oid in lst_oids:
            lst_attachments = ago_flayer.attachments.get_list(oid=oid)
            if lst_attachments:
                original_feature = [f for f in all_features if f.attributes['OBJECTID'] == oid][0]
                try:
                    lst_pictures = original_feature.attributes[fld_picture].split(',')
                except:
                    lst_pictures = []
                lst_new_pictures = [pic for pic in lst_pictures if pic not in lst_os_pictures]
                if not lst_new_pictures:
                    continue

                for attach in lst_attachments:
                    attach_name = attach['name']
                    if attach_name in lst_new_pictures:
                        self.logger.info(f'Copying {attach_name} to object storage')
                        attach_id = attach['id']
                        attach_file = ago_flayer.attachments.download(oid=oid, attachment_id=attach_id)[0]
                        ostore_path = f'{self.bucket_prefix}/{folder}/{attach_name}'

                        self.boto_resource.meta.client.upload_file(attach_file, self.trapper_bucket, ostore_path)


    def create_wild_report(self) -> None:
        self.logger.info('Creating WILD report')

        dict_wild = defaultdict(TrapYear)
        ago_item = self.gis.content.get(self.ago_traps)
        ago_flayer = ago_item.layers[0]

        trap_fset = ago_flayer.query()
        if len(trap_fset.features) == 0:
            return

        trap_sdf = trap_fset.sdf
        trap_sdf.reset_index()

        for index in trap_sdf.index:
            oid = trap_sdf['OBJECTID'][index]
            trapline = trap_sdf['TRAPLINE_ID'][index]
            set_id = trap_sdf['SET_UNIQUE_ID'][index]
            trap_geom = trap_sdf['SHAPE'][index]

            self.logger.info(f'Getting info from {set_id}')

            trap_checks = ago_flayer.query_related_records(object_ids=oid, relationship_id='0')
            rel_groups = trap_checks['relatedRecordGroups']
            if not rel_groups:
                continue

            wmu = self.find_intersect(trap_geom=trap_geom, ago_item_id=self.ago_wmu, field='WILDLIFE_MGMT_UNIT_ID')

            if wmu:
                wmu_split = wmu.split('-')
                wmu_sub = wmu_split[1]
                wmu_sub = wmu_sub if len(wmu_sub) == 2 else f'0{wmu_sub}'
                wmu = f'{wmu_split[0]}{wmu_sub}'

            park = self.find_intersect(trap_geom=trap_geom, ago_item_id=self.ago_parks, field='PROTECTED_LANDS_NAME')
            
            for grp in rel_groups:
                for record in grp['relatedRecords']:
                    # self.logger.info(record)
                    trap_year = self.get_trap_season(trap_date=record['attributes']['CHECK_DATE'])
                    month = dt.datetime.fromtimestamp(int(record['attributes']['CHECK_DATE'])/1000).strftime('%B')
                    trapline_type = 'Registered Trapline' if trapline.lower() != 'unknown' else 'Private Property'
                    species = str(record['attributes']['SPECIES']).title()
                    comments = record['attributes']['CAPTURE_COMMENTS']
                    species = '' if species == 'Na' else species if species !='Other' else comments
                    harvest = 'Yes' if species else 'No'
                    sex = record['attributes']['SEX']
                    m_count = 0 if sex != 'Male' else 1
                    f_count = 0 if sex != 'Female' else 1
                    u_count = 1 if sex == 'NA' and species else 0
                    park_harvest = 'No'
                    permit = ''
                    park_name = ''
                    if park and harvest == 'Yes':
                        park_harvest = 'Yes'
                        park_name = park
                        permit = 'FILL IN WITH PERMIT AUTHORIZATION NUMBER'

                    dict_wild[trap_year].dict_trapline[trapline].trapline_type = trapline_type
                    if dict_wild[trap_year].dict_trapline[trapline].harvest == 'No':
                        dict_wild[trap_year].dict_trapline[trapline].harvest = harvest
                    dict_wild[trap_year].dict_trapline[trapline].dict_wmu[wmu].dict_month[month] \
                            .dict_park[park_name].park_harvest = park_harvest
                    dict_wild[trap_year].dict_trapline[trapline].dict_wmu[wmu].dict_month[month] \
                            .dict_park[park_name].permit = permit
                    dict_wild[trap_year].dict_trapline[trapline].dict_wmu[wmu].dict_month[month] \
                            .dict_park[park_name].dict_species[species].female_count += f_count
                    dict_wild[trap_year].dict_trapline[trapline].dict_wmu[wmu].dict_month[month] \
                            .dict_park[park_name].dict_species[species].male_count += m_count
                    dict_wild[trap_year].dict_trapline[trapline].dict_wmu[wmu].dict_month[month] \
                            .dict_park[park_name].dict_species[species].unknown_count += u_count


        
        
        columns = ['Trapping Licence Year', 'Did Harvest Occur?', 'Trapline Type', 'Trapline Number', 'Month', 
                   'Species', 'WMU', 'Male Count', 'Female Count', 'Unknown Sex Count', 'Harvest in Park?', 
                   'Park Name', 'PERMITAUTHORIZATIONNUMBER']
        for trapyear in sorted(dict_wild.keys()):
            out_dir = os.path.join('wild_reports', trapyear.replace('/', '_'))
            try:
                os.makedirs(out_dir)
            except:
                pass
            self.logger.info(f'Working on trap year: {trapyear}')
            for trapline in sorted(dict_wild[trapyear].dict_trapline.keys()):
                self.logger.info(f'Creating report for {trapline}')
                xl_file = os.path.join(out_dir, f'{trapline.lower()}_wild_report.xlsx')
                lst_traps = []
                trapline_result = dict_wild[trapyear].get_list(trapline=trapline)
                lst_trapline = trapline_result[0]
                catch_count = trapline_result[1]
                if catch_count > 0:
                    for trap in lst_trapline:
                        lst_traps.append([trapyear] + trap)
                else:
                    lst_traps
                self.logger.info(lst_traps)
                df = pd.DataFrame(data=lst_traps, columns=columns)
                sheet_name = trapline
                
                with pd.ExcelWriter(xl_file, date_format='yyyy-mm-dd', datetime_format='yyyy-mm-dd') as xl_writer:
                    df.to_excel(xl_writer, sheet_name=sheet_name, index=False)

                    ws = xl_writer.sheets[sheet_name]

                    dim_holder = DimensionHolder(worksheet=ws)

                    for col in range(ws.min_column, ws.max_column + 1):
                        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=20)

                        ws.column_dimensions = dim_holder

                    for row in ws.iter_rows():
                        for cell in row:
                            cell.alignment = Alignment(wrap_text=True)

                ostore_path = f'{self.bucket_prefix}/{xl_file}'
                self.logger.info('Uploading document to object storage')
                self.boto_resource.meta.client.upload_file(xl_file, self.trapper_bucket, ostore_path)
            
    
    def find_intersect(self, trap_geom, ago_item_id, field):
        ago_item = self.gis.content.get(ago_item_id)
        ago_flayer = ago_item.layers[0]
        ago_feat = ago_flayer.query(where='1=1', geometry_filter=filters.intersects(trap_geom))
        if len(ago_feat.features) == 0:
            return None
        ago_sdf = ago_feat.sdf
        for i in ago_sdf.index:
            val = ago_sdf[field][i]
            break
        return str(val).title()
    
    def get_trap_season(self, trap_date):
        date = dt.datetime.fromtimestamp(int(trap_date)/1000)
        mnth = date.month
        year = date.year

        if mnth >= 7:
            season = f'{year}/{str(year + 1)[-2:]}'
        else:
            season = f'{year-1}/{str(year)[-2:]}'

        return season

    def create_excel(self) -> None:
        self.logger.info('Creating report')
        
        xl_report = 'trapper_data_report.xlsx'
        with pd.ExcelWriter(xl_report, date_format='yyyy-mm-dd', datetime_format='yyyy-mm-dd') as xl_writer:
            self.create_sheet(xl_writer=xl_writer, sheet_name='traps', ago_layer=self.ago_traps, 
                              drop_columns=['GlobalID', 'OBJECTID', 'EDIT_DATE', 'CALCULATE_DATE', 'SHAPE'], date_field='START_DATE')
            
            self.create_sheet(xl_writer=xl_writer, sheet_name='trap checks', ago_layer=self.ago_traps, 
                              drop_columns=['GlobalID', 'OBJECTID', 'EDIT_DATE', 'CALCULATE_DATE', 'TRAPSET_TYPES'], date_field='CHECK_DATE')
            
            self.create_sheet(xl_writer=xl_writer, sheet_name='fisher', ago_layer=self.ago_fisher, 
                              drop_columns=['GlobalID', 'OBJECTID', 'EDIT_DATE', 'CALCULATE_DATE', 'SHAPE'], date_field='OBSERVATION_DATE')

        ostore_path = f'{self.bucket_prefix}/{os.path.basename(xl_report)}'

        self.logger.info('Uploading document to object storage')
        self.boto_resource.meta.client.upload_file(xl_report, self.trapper_bucket, ostore_path)
    

    def create_sheet(self, xl_writer, sheet_name, ago_layer, drop_columns, date_field) -> None:
        self.logger.info(f'Generating {sheet_name} sheet')
        ago_item = self.gis.content.get(ago_layer)
        if sheet_name != 'trap checks':
            ago_flayer = ago_item.layers[0]
        else:
            ago_flayer = ago_item.tables[0]
        ago_fset = ago_flayer.query()
        if len(ago_fset.features) == 0:
            return
        df = ago_fset.sdf
        df.drop(drop_columns, axis=1, inplace=True)
        df[date_field] = pd.to_datetime(df[date_field]).dt.date
        df.to_excel(xl_writer, sheet_name=sheet_name, index=False)

        ws = xl_writer.sheets[sheet_name]

        dim_holder = DimensionHolder(worksheet=ws)

        for col in range(ws.min_column, ws.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=20)

            ws.column_dimensions = dim_holder
        
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

        

class TrapYear:
    def __init__(self) -> None:
        self.dict_trapline = defaultdict(self.Trapline)

    def get_list(self, trapline) -> list:
        lst_traplines = []
        for lst_wmu in self.dict_trapline[trapline].get_list():
            lst_traplines.append([self.dict_trapline[trapline].harvest, self.dict_trapline[trapline].trapline_type, trapline] + lst_wmu)

        return [lst_traplines, self.dict_trapline[trapline].catch_count]


    class Trapline:
        def __init__(self) -> None:
            self.trapline_type = ''
            self.harvest = 'No'
            self.dict_wmu = defaultdict(self.WMU)
            self.catch_count = 0
        
        def get_list(self) -> list:
            lst_wmu = []
            for wmu in self.dict_wmu:
                for lst_month in self.dict_wmu[wmu].get_list():
                    lst_wmu.append(lst_month[:2] + [wmu] + lst_month[2:])
                if self.dict_wmu[wmu].catch_count == 0:
                    lst_wmu = [['',''] + [wmu] + [0, 0, 0, 'No', '', '']]
                self.catch_count += self.dict_wmu[wmu].catch_count

            return lst_wmu

        class WMU:
            def __init__(self) -> None:
                self.dict_month = defaultdict(self.Month)
                self.catch_count = 0
            
            def get_list(self) -> list:
                lst_month = []
                for month in self.dict_month:
                    for lst_park in self.dict_month[month].get_list():
                        lst_month.append([month] + lst_park)
                    self.catch_count += self.dict_month[month].catch_count
                return lst_month


            class Month:
                def __init__(self) -> None:
                    self.dict_park = defaultdict(self.Park)
                    self.catch_count = 0
                
                def get_list(self) -> list:
                    lst_park = []
                    for park in self.dict_park:
                        for lst_species in self.dict_park[park].get_list():
                            lst_park.append(lst_species + [self.dict_park[park].park_harvest, park, 
                                        self.dict_park[park].permit])
                        self.catch_count += self.dict_park[park].catch_count
                    return lst_park

                class Park:
                    def __init__(self) -> None:
                        self.park_harvest = ''
                        self.permit = ''
                        self.dict_species = defaultdict(self.Species)
                        self.catch_count = 0
                    
                    def get_list(self) -> list:
                        lst_species = []
                        for species in self.dict_species:
                            self.catch_count += self.dict_species[species].get_count()
                            if species != '':
                                lst_species.append([species] + self.dict_species[species].get_list())
                        if not lst_species:
                            lst_species = ['', 0, 0, 0]
                        return lst_species
                    
                    class Species:
                        def __init__(self) -> None:
                            self.male_count = 0
                            self.female_count = 0
                            self.unknown_count = 0

                        def get_count(self) -> int:
                            return self.male_count + self.female_count + self.unknown_count

                        def get_list(self) -> list:
                            return [self.male_count, self.female_count, self.unknown_count]




class TrapCheck:
    def __init__(self, trap_year: str='', harvest: str='', tl_type: str='', tl_num: str='', month: str='', 
                 spec: str='', ct_f: int=0, ct_m: int=0, ct_u: int=0, p_harv: str='No', p_name: str='', 
                 permit: str='') -> None:
        self.trap_year = trap_year
        self.harvest = harvest
        self.trapline_type = tl_type
        self.trapline_number = tl_num
        self.month = month
        self.species = spec
        self.male_count = ct_m
        self.female_count = ct_f
        self.unknown_count = ct_u
        self.park_harvest = p_harv
        self.park_name = p_name
        self.permit = permit


def get_col_widths(dataframe):
    # First we find the maximum length of the index column
    idx_max = max([len(str(s)) for s in dataframe.index.values] + [len(str(dataframe.index.name))])
    # Then, we concatenate this to the max of the lengths of column name and its values for each column, left to right

    return [idx_max] + [max([len(str(s)) for s in dataframe[col].values] + [len(col)]) + 4 for col in dataframe.columns]

if __name__ == '__main__':
    run_app()
